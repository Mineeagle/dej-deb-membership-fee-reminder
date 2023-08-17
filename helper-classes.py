import openpyxl
import smtplib
from email.mime.text import MIMEText


class ExcelHandler:
    """Excel handling
    This class allows for primitive access to excel files.
    One file per Instance.

    :param path: Absolute path of the file.

    import openpyxl
    Nothing needs to be changed.
    The standard worksheet will be the active on (last one used).
    """

    def __init__(self, PATH):
        self.PATH = PATH
        self.WB = openpyxl.load_workbook(PATH)
        self.ws = self.WB.active

    def set_worksheet(self, title):
        """Changing of the worksheet
        Since one excel file may contain multiple worksheets, this method
        can be used to change the worksheet.

        :param title: Name of the worksheet.

        The standard worksheet upon instantiation is the active one.
        Nothing needs to be changed.
        """
        self.ws = self.WB[title]

    def read(self, cell):
        """Reading value of cell
        This method returns the value of the given cell.

        read() -> string: Contents of the cell
        :param cell: Input is the cell as a string of type "[col][row]"; for instance "A1"

        Nothing needs to be changed.
        """
        return self.ws[cell].value

    def write(self, cell, data):
        """Write in cell
        This method write data into a cell.

        :param cell: Input is the cell as a string of type "[col][row]"; for instance "A1"
        :param data: The string that should be written into the cell.

        Nothing needs to be changed.
        """
        self.ws[cell] = data
        self.WB.save(self.PATH)

    def get_max_row(self):
        """Get the maximum row used
        This method just returns the last row that is used / has been used.


        get_max_row() -> int: The row number of the last row.

        May be useful for looping.
        Nothing needs to be changed.
        """
        return self.ws.max_row


class EmailSender:
    """Class to send emails
    This class allows for a more easier handling of sending emails.

    :param sender_email: Email of the sender
    :param sender_password: Password for the account of the sender
    :smtp_server: SMTP server; stanard is 'webmail.esperanto.de'
    :smtp_port: port; standard is 587

    Nothing needs to be changed.
    """

    def __init__(
        self,
        sender_email,
        sender_password,
        smtp_server="webmail.esperanto.de",
        smtp_port=587,
    ):
        self.sender_email = sender_email

        self.sender = smtplib.SMTP(smtp_server, smtp_port)
        self.sender.ehlo()
        self.sender.starttls()
        self.sender.ehlo()

        self.sender.login(sender_email, sender_password)

    def send_mail(self, text, subject, recipient):
        """Send a simple email
        This method send a simple email to a recipient.

        send_mail() -> bool: returns True after sending the email
        :param text: body of the email
        :param subject: subject of the email
        :param recipient: address of the recipient

        Nothing needs to be changed.
        """
        mail = MIMEText(text)
        mail["Subject"] = subject
        mail["From"] = self.sender_email
        mail["To"] = recipient

        self.sender.send_message(mail)
        self.sender.close()
        return True
